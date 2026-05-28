import io
import logging
import os
import threading
from collections import Counter
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import HTMLResponse
import openpyxl

from auth import get_request_token

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/excel-mapping", tags=["excel-mapping"])

# ── Config parsing ─────────────────────────────────────────────────────────────

def _parse_sources() -> list[tuple[str, str, int]]:
    """Parse EXCEL_MAPPING_FILE_PATH into [(type, path, 1-based-index)]."""
    raw = os.getenv("EXCEL_MAPPING_FILE_PATH", "").strip()
    if not raw:
        return []
    result = []
    for idx, entry in enumerate(raw.split(","), start=1):
        entry = entry.strip()
        if not entry:
            continue
        type_part, sep, path_part = entry.partition(":")
        if not sep:
            logger.warning(f"Skipping malformed EXCEL_MAPPING_FILE_PATH entry: {entry!r}")
            continue
        type_part = type_part.strip().upper()
        path_part = path_part.strip()
        if type_part not in ("LOCAL", "GCS"):
            logger.warning(f"Unknown source type {type_part!r} in EXCEL_MAPPING_FILE_PATH, skipping.")
            continue
        result.append((type_part, path_part, idx))
    return result

# ── GCS helpers ────────────────────────────────────────────────────────────────

_gcs_client = None
_gcs_init_lock = threading.Lock()


def _get_gcs_client():
    global _gcs_client
    with _gcs_init_lock:
        if _gcs_client is None:
            from google.cloud import storage
            _gcs_client = storage.Client()
    return _gcs_client


def _parse_gcs_path(gcs_url: str) -> tuple[str, str]:
    """gs://bucket/prefix → (bucket, prefix_no_slashes)."""
    without = gcs_url.removeprefix("gs://")
    bucket, _, prefix = without.partition("/")
    return bucket, prefix.strip("/")


def _list_gcs_xlsx(gcs_url: str) -> list[tuple[str, str]]:
    """List .xlsx files (one level deep) at a GCS prefix. Returns [(filename, full_gcs_url)]."""
    bucket_name, prefix = _parse_gcs_path(gcs_url)
    blob_prefix = (prefix + "/") if prefix else ""
    try:
        client = _get_gcs_client()
        result = []
        for blob in client.list_blobs(bucket_name, prefix=blob_prefix):
            rel = blob.name[len(blob_prefix):]
            if not rel or "/" in rel:
                continue  # skip subdirectories
            if rel.endswith(".xlsx"):
                result.append((rel, f"gs://{bucket_name}/{blob.name}"))
        return sorted(result)
    except Exception as e:
        logger.warning(f"Could not list GCS path {gcs_url}: {e}")
        return []


def _download_gcs_blob(gcs_url: str) -> bytes:
    bucket_name, _ = _parse_gcs_path(gcs_url)
    blob_path = gcs_url.removeprefix(f"gs://{bucket_name}/")
    client = _get_gcs_client()
    return client.bucket(bucket_name).blob(blob_path).download_as_bytes()

# ── Local helpers ──────────────────────────────────────────────────────────────

def _list_local_xlsx(path: str) -> list[tuple[str, str]]:
    """List .xlsx files in a local directory. Returns [(filename, full_path)]."""
    try:
        p = Path(path)
        return sorted((f.name, str(f)) for f in p.iterdir() if f.is_file() and f.suffix == ".xlsx")
    except Exception as e:
        logger.warning(f"Could not list local path {path}: {e}")
        return []

# ── File collection & conflict resolution ──────────────────────────────────────

def _collect_files(sources: list[tuple[str, str, int]]) -> list[tuple[str, str, int, str]]:
    """Returns [(filename, src_type, src_idx, full_path)] for all sources in order."""
    result = []
    for src_type, src_path, src_idx in sources:
        files = _list_local_xlsx(src_path) if src_type == "LOCAL" else _list_gcs_xlsx(src_path)
        for filename, full_path in files:
            result.append((filename, src_type, src_idx, full_path))
    return result


def _resolve_names(files: list[tuple[str, str, int, str]]) -> list[tuple[str, str, int, str]]:
    """
    Apply _local<N>/_gcs<N> suffixes when filenames conflict across sources.
    N = 1-based position of the source in the config list.
    Returns [(display_name, src_type, src_idx, full_path)].
    """
    counts = Counter(f[0] for f in files)
    result = []
    for filename, src_type, src_idx, full_path in files:
        base = filename[:-5]  # strip .xlsx
        if counts[filename] > 1:
            suffix = f"_local{src_idx}" if src_type == "LOCAL" else f"_gcs{src_idx}"
            display = base + suffix
        else:
            display = base
        result.append((display, src_type, src_idx, full_path))
    return result

# ── Excel reading ──────────────────────────────────────────────────────────────

def _read_xlsx(src_type: str, full_path: str) -> dict:
    """Skip rows 1–3, row 4 = header, first sheet only. Returns data dict."""
    try:
        if src_type == "LOCAL":
            wb = openpyxl.load_workbook(full_path, read_only=True, data_only=True)
        else:
            wb = openpyxl.load_workbook(io.BytesIO(_download_gcs_blob(full_path)), read_only=True, data_only=True)

        ws = wb.worksheets[0]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(all_rows) < 4:
            return {"error": "File has fewer than 4 rows — empty or corrupt"}

        raw_header = all_rows[3]
        header = [str(h) if h is not None else f"Col{i}" for i, h in enumerate(raw_header, 1)]
        data_rows = all_rows[4:]

        total = len(data_rows)
        mapped = sum(
            1 for r in data_rows
            if r and str(r[0] or "").strip().upper() in ("X", "Y")
        )
        return {
            "error":       None,
            "total_rows":  total,
            "mapped":      mapped,
            "in_progress": total - mapped,
            "header":      header,
            "rows":        [list(r) for r in data_rows],
        }
    except Exception as e:
        logger.error(f"Failed to read {full_path}: {e}")
        return {"error": str(e)}

# ── Cache ──────────────────────────────────────────────────────────────────────

_cache: list[dict] | None = None
_cache_lock = threading.Lock()


def _build_cache() -> list[dict]:
    sources = _parse_sources()
    raw    = _collect_files(sources)
    named  = _resolve_names(raw)
    result = []
    for display_name, src_type, _idx, full_path in named:
        entry = _read_xlsx(src_type, full_path)
        result.append({"display_name": display_name, **entry})
    return result


def _get_cache() -> list[dict]:
    global _cache
    with _cache_lock:
        if _cache is None:
            _cache = _build_cache()
        return list(_cache)


def _invalidate_and_rebuild() -> list[dict]:
    global _cache
    with _cache_lock:
        _cache = _build_cache()
        return list(_cache)

# ── Preview HTML ───────────────────────────────────────────────────────────────

def _build_preview_html(filename: str, header: list[str], rows: list[list]) -> str:
    th_cells = "".join(f"<th>{h}</th>" for h in header)
    body_rows = []
    for i, row in enumerate(rows):
        col1 = str(row[0] if row else "").strip().upper()
        is_mapped = col1 in ("X", "Y")
        cells = []
        for j, cell in enumerate(row):
            val = "" if cell is None else str(cell)
            if j == 0:
                color = "#16a34a" if is_mapped else "#d97706"
                cells.append(f'<td style="color:{color};font-weight:600">{val}</td>')
            else:
                cells.append(f"<td>{val}</td>")
        bg = "#ffffff" if i % 2 == 0 else "#eff6ff"
        body_rows.append(f'<tr style="background:{bg}">{"".join(cells)}</tr>')

    col_count = len(header)
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{filename}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Courier New',monospace;font-size:12px;padding:16px;background:#f8fafc;color:#334155}}
  .toolbar{{display:flex;align-items:center;gap:12px;margin-bottom:10px;flex-wrap:wrap}}
  h1{{font-size:14px;font-weight:700;color:#1e293b}}
  .ctrl{{display:flex;align-items:center;gap:6px;font-size:11px;color:#475569}}
  .ctrl input[type=range]{{width:120px;accent-color:#4472C4}}
  .wrap{{overflow-x:auto}}
  table{{border-collapse:collapse;table-layout:fixed;width:100%;background:#fff;border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,.12)}}
  th{{background:#4472C4;color:#fff;font-weight:700;padding:8px 14px;text-align:left;font-size:11px;white-space:nowrap;position:sticky;top:0;overflow:hidden;text-overflow:ellipsis}}
  td{{padding:6px 14px;border-bottom:1px solid #e2e8f0;color:#000;word-wrap:break-word;overflow-wrap:break-word;overflow:hidden}}
  tr:last-child td{{border-bottom:none}}
</style>
</head>
<body>
<div class="toolbar">
  <h1>{filename}.xlsx</h1>
  <div class="ctrl">
    <span>Col width:</span>
    <input type="range" id="colw" min="60" max="400" value="160" step="10"
           oninput="setColWidth(this.value)">
    <span id="colw-label">160px</span>
  </div>
  <button onclick="setColWidth(60)"  style="font-size:11px;padding:2px 8px;cursor:pointer">Narrow</button>
  <button onclick="setColWidth(160)" style="font-size:11px;padding:2px 8px;cursor:pointer">Default</button>
  <button onclick="setColWidth(300)" style="font-size:11px;padding:2px 8px;cursor:pointer">Wide</button>
</div>
<div class="wrap">
<table id="tbl">
  <colgroup id="cg">{"<col>" * col_count}</colgroup>
  <thead><tr>{th_cells}</tr></thead>
  <tbody>{"".join(body_rows)}</tbody>
</table>
</div>
<script>
function setColWidth(px) {{
  px = Math.max(60, Math.min(400, parseInt(px)));
  document.getElementById('colw').value = px;
  document.getElementById('colw-label').textContent = px + 'px';
  var cols = document.getElementById('cg').getElementsByTagName('col');
  for (var i = 0; i < cols.length; i++) cols[i].style.width = px + 'px';
}}
setColWidth(160);
</script>
</body>
</html>"""

# ── Shared response shaper ─────────────────────────────────────────────────────

def _to_file_list(entries: list[dict]) -> list[dict]:
    return [
        {
            "display_name": e["display_name"],
            "total_rows":   e.get("total_rows"),
            "mapped":       e.get("mapped"),
            "in_progress":  e.get("in_progress"),
            "error":        e.get("error"),
        }
        for e in entries
    ]

# ── Routes ─────────────────────────────────────────────────────────────────────

@router.get("")
async def get_excel_mapping(token: Optional[str] = Depends(get_request_token)):
    """Return cached summary of all Excel mapping files."""
    if not _parse_sources():
        return {"configured": False, "files": []}
    return {"configured": True, "files": _to_file_list(_get_cache())}


@router.post("/refresh")
async def refresh_excel_mapping(token: Optional[str] = Depends(get_request_token)):
    """Bust cache, re-read all files, return fresh summary."""
    if not _parse_sources():
        return {"configured": False, "files": []}
    return {"configured": True, "files": _to_file_list(_invalidate_and_rebuild())}


@router.get("/{filename}/preview", response_class=HTMLResponse)
async def preview_excel_mapping(
    filename: str,
    token: Optional[str] = Depends(get_request_token),
):
    """Return a self-contained HTML page showing the file's first sheet."""
    entries = _get_cache()
    entry = next((e for e in entries if e["display_name"] == filename), None)
    if entry is None:
        raise HTTPException(status_code=404, detail=f"File '{filename}' not found in cache.")
    if entry.get("error"):
        return HTMLResponse(
            f"<html><body style='font-family:sans-serif;padding:2rem'>"
            f"<h2 style='color:#dc2626'>Error loading {filename}.xlsx</h2>"
            f"<p style='color:#6b7280;margin-top:8px'>{entry['error']}</p>"
            f"</body></html>"
        )
    return HTMLResponse(_build_preview_html(filename, entry["header"], entry["rows"]))
