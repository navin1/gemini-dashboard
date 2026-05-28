import io
import json
import logging
import os
import re
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from auth import get_request_token
import bigquery_client

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/schema-audit", tags=["schema-audit"])

# ── Env helpers ────────────────────────────────────────────────────────────────

_ENV_MAP = {
    "dev": ("SCHEMA_AUDIT_DEV_SRC", "SCHEMA_AUDIT_DEV_TGT"),
    "uat": ("SCHEMA_AUDIT_UAT_SRC", "SCHEMA_AUDIT_UAT_TGT"),
    "prd": ("SCHEMA_AUDIT_PRD_SRC", "SCHEMA_AUDIT_PRD_TGT"),
}


def _get_datasets(env: str) -> tuple[str, str]:
    keys = _ENV_MAP.get(env.lower())
    if not keys:
        raise HTTPException(status_code=400, detail=f"Unknown environment: {env}")
    src = os.getenv(keys[0], "").strip()
    tgt = os.getenv(keys[1], "").strip()
    return src, tgt


def _get_src_tbl() -> str:
    return os.getenv("SCHEMA_AUDIT_SRC_TBL", "").strip()


def _get_mysql_info_tbl() -> str:
    return os.getenv("SCHEMA_AUDIT_MYSQL_INFO", "").strip()


_IGNORE_CONFIG_PATH = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "schema_audit_ignore.json")
)


def _merge_ignore(base: dict, override: dict) -> dict:
    """Merge two ignore configs: union tables, union per-table column lists."""
    tables = list(set(base.get("ignore_tables", [])) | set(override.get("ignore_tables", [])))

    def _merge_cols(a: dict, b: dict) -> dict:
        keys = set(a) | set(b)
        return {k: list(set(a.get(k, [])) | set(b.get(k, []))) for k in keys}

    return {
        "ignore_tables":      tables,
        "ignore_src_columns": _merge_cols(base.get("ignore_src_columns", {}), override.get("ignore_src_columns", {})),
        "ignore_tgt_columns": _merge_cols(base.get("ignore_tgt_columns", {}), override.get("ignore_tgt_columns", {})),
    }


def _load_ignore_config(env: str) -> dict:
    """Returns merged all_env + env-specific ignore config. Empty dict on any error or missing key."""
    try:
        with open(_IGNORE_CONFIG_PATH) as f:
            raw = json.load(f)
        all_env = raw.get("all_env", {})
        env_cfg = raw.get(env.lower(), {})
        return _merge_ignore(all_env, env_cfg)
    except FileNotFoundError:
        return {}
    except Exception as e:
        logger.warning(f"Could not load schema_audit_ignore.json: {e}")
        return {}


# ── BigQuery helpers ───────────────────────────────────────────────────────────

def _fetch_table_list(token) -> list[str]:
    src_tbl = _get_src_tbl()
    if not src_tbl:
        return []
    sql = f"SELECT DISTINCT table_name FROM `{src_tbl}` WHERE is_streamed = 1 ORDER BY table_name"
    try:
        rows = bigquery_client.run_query(sql, token)
        return [r["table_name"] for r in rows if r.get("table_name")]
    except Exception as e:
        logger.error(f"Failed to fetch table list: {e}")
        return []


def _sanitize_table_names(names: list[str]) -> list[str]:
    """Allow only safe characters in table names to prevent SQL injection."""
    return [n for n in names if re.match(r"^[A-Za-z0-9_\-]+$", n)]


def _fetch_mysql_src_columns(table_names: list[str], token) -> dict[str, list[dict]] | None:
    """
    For DEV: query SCHEMA_AUDIT_MYSQL_INFO directly (already has column metadata).
    Returns {table_name: [columns]} or None if not configured.
    """
    mysql_info_tbl = _get_mysql_info_tbl()
    if not mysql_info_tbl or not table_names:
        return None

    safe_names = _sanitize_table_names(table_names)
    if not safe_names:
        return {}

    quoted = ", ".join(f"'{n}'" for n in safe_names)
    sql = f"""
        SELECT table_name, column_name, data_type, CAST(ordinal_position AS INT64) AS ordinal_position
        FROM `{mysql_info_tbl}`
        WHERE table_name IN ({quoted})
        ORDER BY table_name, ordinal_position
    """
    try:
        rows = bigquery_client.run_query(sql, token)
    except Exception as e:
        logger.error(f"MYSQL_INFO query failed for {mysql_info_tbl}: {e}")
        return None

    result: dict[str, list[dict]] = {}
    for row in rows:
        tbl = row["table_name"]
        result.setdefault(tbl, []).append({
            "column_name":      row["column_name"],
            "data_type":        row["data_type"],
            "ordinal_position": int(row["ordinal_position"]),
        })
    return result


def _fetch_all_columns(dataset: str, table_names: list[str], token) -> dict[str, list[dict]] | None:
    """
    Returns {table_name: [columns]} for all tables found in the dataset.
    Returns None if dataset is not configured.
    Tables missing from the result simply don't exist in this dataset.
    """
    if not dataset or not table_names:
        return None

    safe_names = _sanitize_table_names(table_names)
    if not safe_names:
        return {}

    quoted = ", ".join(f"'{n}'" for n in safe_names)
    sql = f"""
        SELECT table_name, column_name, data_type, ordinal_position
        FROM `{dataset}`.INFORMATION_SCHEMA.COLUMNS
        WHERE table_name IN ({quoted})
        ORDER BY table_name, ordinal_position
    """
    try:
        rows = bigquery_client.run_query(sql, token)
    except Exception as e:
        logger.error(f"INFORMATION_SCHEMA query failed for {dataset}: {e}")
        return None

    result: dict[str, list[dict]] = {}
    for row in rows:
        tbl = row["table_name"]
        result.setdefault(tbl, []).append({
            "column_name":      row["column_name"],
            "data_type":        row["data_type"],
            "ordinal_position": int(row["ordinal_position"]),
        })
    return result


# ── Schema comparison ──────────────────────────────────────────────────────────

def _describe_mismatch(type_diff: bool, pos_diff: bool) -> str:
    if type_diff and pos_diff:
        return "Type + Position Mismatch"
    if type_diff:
        return "Type Mismatch"
    if pos_diff:
        return "Position Mismatch"
    return ""


def _compare_table(
    table_name: str,
    src_cols: list[dict] | None,
    tgt_cols: list[dict] | None,
) -> tuple[dict, list[dict]]:
    """
    Returns:
      summary — one-row widget dict
      detail  — per-column dicts for Excel
    """
    src_missing = src_cols is None
    tgt_missing = tgt_cols is None

    src_list = src_cols or []
    tgt_list = tgt_cols or []

    src_map = {c["column_name"]: c for c in src_list}
    tgt_map = {c["column_name"]: c for c in tgt_list}

    col_count_diff = len(src_list) - len(tgt_list)

    detail: list[dict] = []
    col_name_mismatches = type_mismatches = pos_mismatches = 0

    # Columns present in SRC
    for col_name, sc in src_map.items():
        if col_name not in tgt_map:
            col_name_mismatches += 1
            detail.append({
                "src_col_name": col_name,
                "src_position": sc["ordinal_position"],
                "tgt_position": "—",
                "src_data_type": sc["data_type"],
                "tgt_data_type": "—",
                "description": "Source Only",
                "status": "Mismatch",
                "_sort": sc["ordinal_position"],
            })
        else:
            tc = tgt_map[col_name]
            type_diff = sc["data_type"].upper() != tc["data_type"].upper()
            pos_diff  = sc["ordinal_position"] != tc["ordinal_position"]
            if type_diff:
                type_mismatches += 1
            if pos_diff:
                pos_mismatches += 1
            desc   = _describe_mismatch(type_diff, pos_diff)
            status = "Mismatch" if (type_diff or pos_diff) else "Match"
            detail.append({
                "src_col_name": col_name,
                "src_position": sc["ordinal_position"],
                "tgt_position": tc["ordinal_position"],
                "src_data_type": sc["data_type"],
                "tgt_data_type": tc["data_type"],
                "description": desc,
                "status": status,
                "_sort": sc["ordinal_position"],
            })

    # Columns present only in TGT
    for col_name, tc in tgt_map.items():
        if col_name not in src_map:
            col_name_mismatches += 1
            detail.append({
                "src_col_name": "—",
                "src_position": "—",
                "tgt_position": tc["ordinal_position"],
                "src_data_type": "—",
                "tgt_data_type": tc["data_type"],
                "description": f"Target Only: {col_name}",
                "status": "Mismatch",
                "_sort": 999_999,  # push TGT-only rows to the end
            })

    detail.sort(key=lambda r: r["_sort"])

    has_mismatch = bool(col_name_mismatches or type_mismatches or pos_mismatches)

    summary = {
        "table_name":          table_name,
        "col_count_diff":      col_count_diff,
        "col_name_mismatches": col_name_mismatches,
        "type_mismatches":     type_mismatches,
        "pos_mismatches":      pos_mismatches,
        "has_mismatch":        has_mismatch,
        "src_missing":         src_missing,
        "tgt_missing":         tgt_missing,
    }
    return summary, detail


def _run_comparison(env: str, token) -> tuple[list[dict], dict[str, list[dict]]]:
    """
    Returns:
      summaries      — sorted list of summary dicts (mismatched first, then alphabetical)
      detail_by_tbl  — {table_name: [detail_rows]}
    """
    ignore          = _load_ignore_config(env)
    ignore_tables   = set(ignore.get("ignore_tables", []))
    ignore_src_cols = ignore.get("ignore_src_columns", {})
    ignore_tgt_cols = ignore.get("ignore_tgt_columns", {})

    _src_ds, tgt_ds = _get_datasets(env)
    table_names     = [t for t in _fetch_table_list(token) if t not in ignore_tables]

    if env.lower() == "dev":
        # PARKED: src_all = _fetch_all_columns(src_ds, table_names, token)
        src_all = _fetch_mysql_src_columns(table_names, token)
    else:
        src_all = _fetch_all_columns(_src_ds, table_names, token)
    tgt_all = _fetch_all_columns(tgt_ds, table_names, token)

    summaries: list[dict] = []
    detail_by_tbl: dict[str, list[dict]] = {}

    for tbl in table_names:
        src_cols = src_all.get(tbl) if src_all is not None else None
        tgt_cols = tgt_all.get(tbl) if tgt_all is not None else None

        if src_cols is not None and tbl in ignore_src_cols:
            skip = {c.lower() for c in ignore_src_cols[tbl]}
            src_cols = [c for c in src_cols if c["column_name"].lower() not in skip]

        if tgt_cols is not None and tbl in ignore_tgt_cols:
            skip = {c.lower() for c in ignore_tgt_cols[tbl]}
            tgt_cols = [c for c in tgt_cols if c["column_name"].lower() not in skip]

        summary, detail = _compare_table(tbl, src_cols, tgt_cols)
        summaries.append(summary)
        detail_by_tbl[tbl] = detail

    # Sort: mismatched tables first (alphabetically), then matched (alphabetically)
    summaries.sort(key=lambda r: (0 if r["has_mismatch"] else 1, r["table_name"]))
    return summaries, detail_by_tbl


# ── Excel builder ──────────────────────────────────────────────────────────────

_BLUE_FILL  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_WHITE_BOLD = Font(name="Courier New", size=11, bold=True, color="FFFFFF")
_DATA_FONT  = Font(name="Courier New", size=11)
_CENTER     = Alignment(horizontal="center", vertical="center")
_LEFT       = Alignment(horizontal="left",   vertical="center")

_TAB_BLUE   = "4472C4"   # Summary sheet
_TAB_GREEN  = "70AD47"   # No mismatches
_TAB_ORANGE = "ED7D31"   # Has mismatches (table present on both sides)
_TAB_RED    = "C00000"   # Entire table missing from SRC or TGT


def _tab_color(summary: dict) -> str:
    if summary["src_missing"] or summary["tgt_missing"]:
        return _TAB_RED
    if summary["has_mismatch"]:
        return _TAB_ORANGE
    return _TAB_GREEN


def _style_header(ws, headers: list[str], col_widths: list[int]) -> None:
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = _BLUE_FILL
        cell.font      = _WHITE_BOLD
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 18


def _write_data_row(ws, row_idx: int, values: list) -> None:
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font      = _DATA_FONT
        cell.alignment = _LEFT


def _safe_sheet_name(name: str) -> str:
    """Truncate to 31 chars and strip characters invalid in Excel sheet names."""
    invalid = r'\/?*[]:'
    for ch in invalid:
        name = name.replace(ch, "_")
    return name[:31]


def _build_excel(summaries: list[dict], detail_by_tbl: dict[str, list[dict]]) -> bytes:
    wb = Workbook()

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_properties.tabColor = _TAB_BLUE

    sum_headers = [
        "Table Name", "SRC Column Name", "SRC Position",
        "TGT Position", "SRC Data Type", "TGT Data Type",
        "Description", "Status",
    ]
    sum_widths = [30, 35, 16, 16, 25, 25, 38, 12]
    _style_header(ws_sum, sum_headers, sum_widths)

    row_idx = 2
    for summary in summaries:
        tbl  = summary["table_name"]
        rows = detail_by_tbl.get(tbl, [])
        for dr in rows:
            if dr["status"] == "Mismatch":
                _write_data_row(ws_sum, row_idx, [
                    tbl,
                    dr["src_col_name"],
                    dr["src_position"],
                    dr["tgt_position"],
                    dr["src_data_type"],
                    dr["tgt_data_type"],
                    dr["description"],
                    dr["status"],
                ])
                row_idx += 1

    ws_sum.freeze_panes = "A2"

    # ── Per-table sheets ─────────────────────────────────────────────────────
    tbl_headers = [
        "SRC Column Name", "SRC Position", "TGT Position",
        "SRC Data Type", "TGT Data Type", "Description", "Status",
    ]
    tbl_widths = [35, 16, 16, 25, 25, 38, 12]

    # Iterate summaries so order matches summary sort (mismatched first)
    for summary in summaries:
        tbl  = summary["table_name"]
        rows = detail_by_tbl.get(tbl, [])

        ws = wb.create_sheet(title=_safe_sheet_name(tbl))
        ws.sheet_properties.tabColor = _tab_color(summary)
        _style_header(ws, tbl_headers, tbl_widths)

        for r_idx, dr in enumerate(rows, start=2):
            _write_data_row(ws, r_idx, [
                dr["src_col_name"],
                dr["src_position"],
                dr["tgt_position"],
                dr["src_data_type"],
                dr["tgt_data_type"],
                dr["description"],
                dr["status"],
            ])
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Routes ─────────────────────────────────────────────────────────────────────

@router.get("/{env}")
async def get_schema_audit(
    env: str,
    token: Optional[str] = Depends(get_request_token),
):
    """Widget summary endpoint — one row per table."""
    _src_ds, tgt_ds = _get_datasets(env)
    if env.lower() == "dev":
        configured = bool(_get_src_tbl() and _get_mysql_info_tbl() and tgt_ds)
    else:
        configured = bool(_get_src_tbl() and _src_ds and tgt_ds)
    if not configured:
        return {"configured": False, "tables": []}

    try:
        summaries, _ = _run_comparison(env, token)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Schema audit failed for env={env}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    return {"configured": True, "tables": summaries}


@router.get("/{env}/download")
async def download_schema_audit(
    env: str,
    token: Optional[str] = Depends(get_request_token),
):
    """Excel download endpoint."""
    if not _get_src_tbl():
        raise HTTPException(status_code=503, detail="SCHEMA_AUDIT_SRC_TBL is not configured.")
    _src_ds, tgt_ds = _get_datasets(env)
    if env.lower() == "dev":
        if not _get_mysql_info_tbl() or not tgt_ds:
            raise HTTPException(status_code=503, detail="SCHEMA_AUDIT_MYSQL_INFO / SCHEMA_AUDIT_DEV_TGT not configured.")
    else:
        if not _src_ds or not tgt_ds:
            raise HTTPException(status_code=503, detail=f"SCHEMA_AUDIT_{env.upper()}_SRC / _TGT not configured.")

    try:
        summaries, detail_by_tbl = _run_comparison(env, token)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Schema audit download failed for env={env}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    excel_bytes = _build_excel(summaries, detail_by_tbl)
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename    = f"schema_mismatch_{env}_{timestamp}.xlsx"

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
